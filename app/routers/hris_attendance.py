"""
GPA-ERP HRIS — Phase H2: Absensi & Cuti
Endpoints for attendance (with geolocation + face verification) and leave management.
"""
from __future__ import annotations

import csv
import io
import logging
import math
import time
import uuid
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Annotated

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_

from app.audit import model_to_dict, write_audit
from app.config import get_settings
from app.database import get_db
from app.dependencies import CurrentUser, get_client_ip, require_role
from app.hris_access import ensure_employee_can_use_self_service
from app.hris_time import (
    MAX_BROWSER_TIMEZONE_OFFSET,
    MIN_BROWSER_TIMEZONE_OFFSET,
    local_date_for_employee,
    timezone_from_browser_offset,
)
from app.models import (
    AttendanceRecord, AttendanceSource,
    Employee, LeaveBalance, LeaveCategory, LeaveRequest, LeaveRequestStatus, LeaveType,
    RoleName, WorkGroup, WorkLocation, WorkLocationType, effective_roles,
    OvertimeRequest, OvertimeRequestStatus, HolidayCalendar,
)
from app.menu_permissions import require_menu_access
from app.notify import push, push_to_role
from app.schemas import (
    AttendanceManualCreate, AttendanceRecordResponse, AttendanceSummaryItem,
    LeaveActionRequest, LeaveBalanceResponse, LeaveRequestCreate,
    LeaveRequestResponse, LeaveTypeCreate, LeaveTypeResponse,
    MessageResponse, PaginatedResponse,
    WorkLocationCreate, WorkLocationResponse, WorkLocationUpdate,
    OvertimeRequestCreate, OvertimeRequestResponse, OvertimeActionRequest,
    LeaveCalendarItem,
)

router = APIRouter(prefix="/hris", tags=["HRIS – Attendance & Leave"])

_hr_roles  = (RoleName.SUPER_ADMIN, RoleName.MD, RoleName.GA, RoleName.HR)
_mgr_roles = (RoleName.SUPER_ADMIN, RoleName.MD, RoleName.PM, RoleName.PROJECT_CONTROL, RoleName.GA, RoleName.HR)

_UPLOAD_ROOT = Path(get_settings().UPLOAD_DIR)
_SELFIE_DIR = _UPLOAD_ROOT / "selfies"
_SELFIE_DIR.mkdir(parents=True, exist_ok=True)
_SELFIE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
_SELFIE_CONTENT_TYPES = {"image/jpeg", "image/png"}
_MAX_SELFIE_BYTES = 5 * 1024 * 1024

_LEAVE_CERT_DIR = _UPLOAD_ROOT / "leave_certificates"
_LEAVE_CERT_DIR.mkdir(parents=True, exist_ok=True)
_LEAVE_CERT_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png"}
_LEAVE_CERT_CONTENT_TYPES = {"application/pdf", "image/jpeg", "image/png"}
_MAX_LEAVE_CERT_BYTES = 10 * 1024 * 1024
_ORPHAN_CERT_MAX_AGE_SECONDS = 24 * 60 * 60

# Default leave approval chain: GA reviews, MD approves
_LEAVE_APPROVAL_CHAIN = ["GA", "MD"]


def _cleanup_orphan_leave_certificates(db: Session, user_id: int) -> int:
    """Remove stale certificates that were never attached to a leave request."""
    removed = 0
    cutoff = time.time() - _ORPHAN_CERT_MAX_AGE_SECONDS
    for path in _LEAVE_CERT_DIR.glob(f"user_{user_id}_*"):
        file_url = f"/uploads/leave_certificates/{path.name}"
        linked = db.query(LeaveRequest.id).filter(
            LeaveRequest.doctor_cert_url == file_url,
        ).first()
        if not linked and path.is_file() and path.stat().st_mtime < cutoff:
            path.unlink(missing_ok=True)
            removed += 1
    return removed


# ─── Overtime calculation (Permenaker No. 2 Tahun 2023) ─────────────────────

def _calculate_overtime(
    clock_in:  datetime,
    clock_out: datetime,
    is_weekend: bool,
    is_holiday: bool,
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """
    Returns (hours_regular, hours_ot_weekday, hours_ot_weekend, hours_ot_holiday).
    Permenaker 2023 rules:
      Weekday: first 8h regular, OT starts after 8h
      Weekend/Holiday: all hours are OT
    """
    MAX_DAILY_HOURS = Decimal("24")  # sanity cap — forgot to clock out guard
    total_hours = max(Decimal(0), min(
        Decimal(str((clock_out - clock_in).total_seconds() / 3600)),
        MAX_DAILY_HOURS,
    ))
    REGULAR_HOURS = Decimal("8")

    if is_holiday:
        return Decimal("0"), Decimal("0"), Decimal("0"), total_hours
    if is_weekend:
        return Decimal("0"), Decimal("0"), total_hours, Decimal("0")

    # Weekday
    regular = min(total_hours, REGULAR_HOURS)
    ot      = max(Decimal("0"), total_hours - REGULAR_HOURS)
    return regular, ot, Decimal("0"), Decimal("0")


def _is_weekend(d: date) -> bool:
    return d.weekday() >= 5  # Saturday=5, Sunday=6


def _is_holiday(db: Session, attendance_date: date) -> bool:
    return db.query(HolidayCalendar.id).filter(HolidayCalendar.date == attendance_date).first() is not None


def _leave_duration(db: Session, start_date: date, end_date: date) -> tuple[int, list[dict]]:
    """Count Monday-Friday leave days, excluding configured holidays."""
    if end_date < start_date:
        raise HTTPException(422, "end_date must be greater than or equal to start_date")

    holidays = (
        db.query(HolidayCalendar)
        .filter(HolidayCalendar.date >= start_date, HolidayCalendar.date <= end_date)
        .all()
    )
    holiday_by_date = {holiday.date: holiday for holiday in holidays}
    excluded_holidays: list[dict] = []
    working_days = 0

    for offset in range((end_date - start_date).days + 1):
        current_date = start_date + timedelta(days=offset)
        if _is_weekend(current_date):
            continue
        holiday = holiday_by_date.get(current_date)
        if holiday:
            excluded_holidays.append({
                "date": holiday.date.isoformat(),
                "name": holiday.name,
            })
            continue
        working_days += 1

    return working_days, excluded_holidays


def _leave_days_by_year(db: Session, start_date: date, end_date: date) -> dict[int, int]:
    allocations: dict[int, int] = {}
    for year in range(start_date.year, end_date.year + 1):
        segment_start = max(start_date, date(year, 1, 1))
        segment_end = min(end_date, date(year, 12, 31))
        days, _ = _leave_duration(db, segment_start, segment_end)
        if days:
            allocations[year] = days
    return allocations


def _get_or_create_leave_balance(
    db: Session,
    employee_id: int,
    leave_type: LeaveType,
    year: int,
    lock: bool = False,
) -> LeaveBalance:
    query = db.query(LeaveBalance).filter(
        LeaveBalance.employee_id == employee_id,
        LeaveBalance.leave_type_id == leave_type.id,
        LeaveBalance.year == year,
    )
    if lock:
        query = query.with_for_update()
    balance = query.first()
    if balance is None:
        balance = LeaveBalance(
            employee_id=employee_id,
            leave_type_id=leave_type.id,
            year=year,
            accrued=leave_type.max_days_per_year or 0,
            used=0,
        )
        db.add(balance)
        db.flush()
    return balance


def _check_or_deduct_leave_balances(
    db: Session,
    employee_id: int,
    leave_type: LeaveType,
    allocations: dict[int, int],
    deduct: bool = False,
) -> None:
    if leave_type.category in (LeaveCategory.MATERNITY, LeaveCategory.PATERNITY):
        return
    if leave_type.max_days_per_year is None:
        return

    balances: list[tuple[LeaveBalance, int]] = []
    for year, days in allocations.items():
        balance = _get_or_create_leave_balance(
            db, employee_id, leave_type, year, lock=deduct,
        )
        if balance.remaining < days:
            raise HTTPException(
                422,
                f"Insufficient {leave_type.name} balance for {year}: "
                f"{balance.remaining} days remaining, {days} required",
            )
        balances.append((balance, days))
    if deduct:
        for balance, days in balances:
            balance.used += days


def _link_overtime_requests_to_attendance(db: Session, record: AttendanceRecord) -> list[int]:
    """Attach overtime requests for the same employee/date to an attendance record."""
    requests = (
        db.query(OvertimeRequest)
        .filter(
            OvertimeRequest.employee_id == record.employee_id,
            OvertimeRequest.date == record.date,
            OvertimeRequest.attendance_id.is_(None),
        )
        .all()
    )
    for overtime_request in requests:
        overtime_request.attendance_id = record.id
    return [overtime_request.id for overtime_request in requests]


def _require_current_leave_approver(request: LeaveRequest, current_user) -> None:
    if current_user.role.name == RoleName.SUPER_ADMIN or not request.current_approver_role:
        return
    try:
        expected_role = RoleName(request.current_approver_role)
    except ValueError:
        raise HTTPException(409, "Leave request has an invalid approver role")
    if expected_role not in effective_roles(current_user.role.name):
        raise HTTPException(403, f"Approval expected from role: {request.current_approver_role}")


def _visible_employee_ids(db: Session, current_user) -> list[int] | None:
    """None means unrestricted HR access; other roles are scoped to self/team."""
    roles = set(effective_roles(current_user.role.name))
    if roles.intersection({RoleName.SUPER_ADMIN, RoleName.MD, RoleName.GA}):
        return None
    own = db.query(Employee).filter(Employee.user_id == current_user.id).first()
    if not own:
        return []
    if RoleName.PM in roles and own.work_group_id:
        return [
            row[0] for row in db.query(Employee.id)
            .filter(Employee.work_group_id == own.work_group_id)
            .all()
        ]
    return [own.id]


# ─── Geolocation helpers ──────────────────────────────────────────────────────

def _haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Return great-circle distance in metres between two GPS points."""
    R = 6_371_000  # Earth radius in metres
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi       = math.radians(lat2 - lat1)
    dlambda    = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def _check_location(
    db: Session,
    latitude: float,
    longitude: float,
    assigned_location: "WorkLocation | None" = None,
) -> tuple[bool, "WorkLocation | None", float]:
    """
    Check GPS coords against WorkLocations.

    If the employee has an assigned work_location, validate ONLY against that
    location — so a Jakarta employee can't clock-in at a Berau site and vice versa.

    If no location is assigned, fall back to checking all active locations
    (previous behaviour — matches the nearest one within any radius).

    The caller decides whether a failed geofence check should block clock-in.
    """
    if assigned_location is not None:
        # Validate strictly against the employee's assigned location only
        locations = [assigned_location]
    else:
        locations = db.query(WorkLocation).filter(WorkLocation.is_active == True).all()

    best: WorkLocation | None = None
    best_dist = float("inf")
    for loc in locations:
        dist = _haversine(latitude, longitude, float(loc.latitude), float(loc.longitude))
        if dist <= loc.radius_meters and dist < best_dist:
            best, best_dist = loc, dist
    # If no match found, compute distance to nearest for reporting
    if best is None:
        nearest_dist = float("inf")
        for loc in locations:
            dist = _haversine(latitude, longitude, float(loc.latitude), float(loc.longitude))
            if dist < nearest_dist:
                nearest_dist = dist
        return False, None, nearest_dist if nearest_dist != float("inf") else 0.0
    return True, best, best_dist


# ─── Attendance: clock-in (mobile, geolocation + selfie) ─────────────────────

@router.post("/attendance/clock-in", response_model=AttendanceRecordResponse,
             summary="Mobile clock-in with GPS + selfie",
             dependencies=[Depends(require_menu_access("hris_attendance"))])
async def clock_in(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    latitude:     float | None        = Form(None),
    longitude:    float | None        = Form(None),
    accuracy:     float | None        = Form(None),
    employee_id:  int | None         = Form(None),
    timezone_offset_minutes: int      = Form(
        ...,
        ge=MIN_BROWSER_TIMEZONE_OFFSET,
        le=MAX_BROWSER_TIMEZONE_OFFSET,
    ),
    selfie:       UploadFile | None  = File(None),
    note:          str | None        = Form(None),
):
    """
    Mobile clock-in: accepts GPS coordinates + selfie photo.
    - If employee_id not provided, uses the employee linked to current_user.
    - Runs face verification if employee has a registered face embedding.
    - Creates or updates the AttendanceRecord for today.
    """
    # Resolve employee
    if employee_id:
        emp = db.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise HTTPException(404, "Employee not found")
        # Only HR/admin can clock in for others
        if emp.user_id != current_user.id and current_user.role.name not in _hr_roles:
            raise HTTPException(403, "Can only clock in for yourself")
    else:
        emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not emp:
            raise HTTPException(404, "No employee record linked to your account")
    ensure_employee_can_use_self_service(emp)

    if latitude is None or longitude is None:
        raise HTTPException(422, "GPS coordinates are required for mobile clock-in")
    if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
        raise HTTPException(422, "Invalid GPS coordinates")
    if accuracy is not None and accuracy < 0:
        raise HTTPException(422, "GPS accuracy cannot be negative")
    if selfie is None:
        raise HTTPException(422, "A selfie is required for mobile clock-in")

    today      = local_date_for_employee(emp, timezone_offset_minutes)
    now        = datetime.now(timezone.utc)
    face_detected: bool             = False
    face_confidence: Decimal | None = None
    selfie_url: str | None          = None
    location_ok: bool | None        = None
    location_distance_m: Decimal | None = None

    open_record = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.employee_id == emp.id,
            AttendanceRecord.clock_in.isnot(None),
            AttendanceRecord.clock_out.is_(None),
        )
        .order_by(AttendanceRecord.date.desc())
        .first()
    )
    if open_record:
        if open_record.date != today:
            raise HTTPException(
                409,
                f"Clock out the open attendance record from {open_record.date} before starting a new shift",
            )
        raise HTTPException(409, "Already clocked in today")

    record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == emp.id,
        AttendanceRecord.date == today,
    ).first()
    if record and record.clock_out:
        raise HTTPException(409, "Attendance for today is already completed")

    # Mobile attendance requires an active assigned work location and geofence.
    assigned_wl: WorkLocation | None = emp.work_location if emp.work_location_id else None
    if assigned_wl is None or not assigned_wl.is_active:
        raise HTTPException(409, "An active work location must be assigned before mobile clock-in")
    matched_loc: WorkLocation | None = None
    ok, matched_loc, dist = _check_location(db, latitude, longitude, assigned_wl)
    location_ok = ok
    location_distance_m = Decimal(str(round(dist, 1)))
    if not ok:
        raise HTTPException(
            422,
            f"Clock-in is outside {assigned_wl.name} geofence "
            f"({dist:.0f} m; allowed radius {assigned_wl.radius_meters} m)",
        )

    # Process selfie — detect whether a face is present (no identity matching)
    if selfie:
        ext = Path(selfie.filename or "").suffix.lower()
        if ext not in _SELFIE_EXTENSIONS or selfie.content_type not in _SELFIE_CONTENT_TYPES:
            raise HTTPException(400, "Selfie must be a JPEG or PNG image")
        selfie_bytes = await selfie.read()
        if not selfie_bytes:
            raise HTTPException(400, "Selfie file is empty")
        if len(selfie_bytes) > _MAX_SELFIE_BYTES:
            raise HTTPException(413, "Selfie exceeds the 5 MB limit")
        filename = f"{emp.id}_{today.isoformat()}_{uuid.uuid4().hex[:8]}{ext}"
        dest     = _SELFIE_DIR / filename
        dest.write_bytes(selfie_bytes)
        selfie_url = f"/uploads/selfies/{filename}"

        try:
            from app.hris_face import detect_face
            face_detected, conf = detect_face(selfie_bytes)
            face_confidence = Decimal(str(conf))
        except Exception as exc:
            logger.warning("Face detection error during clock-in: %s", exc)
            # Don't block clock-in if detection fails

    if record:
        if record.clock_in is None:
            record.clock_in = now
        if selfie_url:
            record.selfie_url      = selfie_url
            record.face_verified   = face_detected
            record.face_confidence = face_confidence
        if latitude is not None:
            record.latitude                  = Decimal(str(latitude))
            record.longitude                 = Decimal(str(longitude))
            record.accuracy                  = Decimal(str(accuracy)) if accuracy is not None else None
            record.location_ok               = location_ok
            record.location_distance_m       = location_distance_m
            record.matched_work_location_id  = matched_loc.id if matched_loc else None
        if note is not None:
            record.note = note
    else:
        record = AttendanceRecord(
            employee_id              = emp.id,
            date                     = today,
            clock_in                 = now,
            source                   = AttendanceSource.MOBILE,
            latitude                 = Decimal(str(latitude)) if latitude is not None else None,
            longitude                = Decimal(str(longitude)) if longitude is not None else None,
            accuracy                 = Decimal(str(accuracy)) if accuracy is not None else None,
            location_ok              = location_ok,
            location_distance_m      = location_distance_m,
            matched_work_location_id = matched_loc.id if matched_loc else None,
            selfie_url               = selfie_url,
            face_verified            = face_detected,
            face_confidence          = face_confidence,
            note                     = note,
        )
        db.add(record)

    db.flush()
    linked_overtime_ids = _link_overtime_requests_to_attendance(db, record)
    write_audit(db, "AttendanceRecord", record.id, "CLOCK_IN",
                changed_by=current_user.id,
                after={"employee_id": emp.id, "date": str(today),
                       "face_detected": face_detected,
                       "face_confidence": str(face_confidence or ""),
                       "has_selfie": selfie_url is not None,
                       "location_ok": location_ok,
                       "location_distance_m": str(location_distance_m or ""),
                       "linked_overtime_request_ids": linked_overtime_ids})
    db.commit()
    db.refresh(record)

    # Flag to HR if selfie was submitted but no face was detected
    if selfie and not face_detected:
        push_to_role(db, RoleName.GA,
                     "Absensi: Wajah Tidak Terdeteksi",
                     f"{emp.full_name} clock-in namun tidak ada wajah pada selfie",
                     "/hris/attendance")
        db.commit()

    return record


# ─── Attendance: clock-out ────────────────────────────────────────────────────

@router.post("/attendance/clock-out", response_model=AttendanceRecordResponse,
             summary="Clock out — calculates hours worked",
             dependencies=[Depends(require_menu_access("hris_attendance"))])
def clock_out(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    employee_id:  int | None = Query(None),
    note:         str | None = Query(None),
    timezone_offset_minutes: int = Query(
        ...,
        ge=MIN_BROWSER_TIMEZONE_OFFSET,
        le=MAX_BROWSER_TIMEZONE_OFFSET,
    ),
):
    if employee_id:
        emp = db.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise HTTPException(404, "Employee not found")
        if emp.user_id != current_user.id and current_user.role.name not in _hr_roles:
            raise HTTPException(403, "Can only clock out for yourself")
    else:
        emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not emp:
            raise HTTPException(404, "No employee record linked to your account")
    ensure_employee_can_use_self_service(emp)

    today      = local_date_for_employee(emp, timezone_offset_minutes)
    now        = datetime.now(timezone.utc)

    record = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.employee_id == emp.id,
            AttendanceRecord.clock_in.isnot(None),
            AttendanceRecord.clock_out.is_(None),
        )
        .order_by(AttendanceRecord.date.desc())
        .first()
    )

    if not record:
        raise HTTPException(409, "No open clock-in found")
    if record.clock_out:
        raise HTTPException(409, "Already clocked out today")

    record.clock_out = now
    is_holiday = _is_holiday(db, record.date)
    if record.clock_in:
        weekend = _is_weekend(record.date)
        reg, ot_wd, ot_we, ot_hol = _calculate_overtime(
            record.clock_in, now, weekend, is_holiday
        )
        record.hours_regular          = reg
        record.hours_overtime_weekday = ot_wd
        record.hours_overtime_weekend = ot_we
        record.hours_overtime_holiday = ot_hol

    if note:
        record.note = note

    write_audit(db, "AttendanceRecord", record.id, "CLOCK_OUT",
                changed_by=current_user.id,
                after={"clock_out": str(now),
                       "hours_regular": str(record.hours_regular or ""),
                       "is_holiday": is_holiday})
    db.commit()
    db.refresh(record)
    return record


# ─── Attendance: manual entry (HR admin) ─────────────────────────────────────

@router.post("/attendance", response_model=AttendanceRecordResponse, status_code=201,
             summary="Manual attendance entry (HR admin)",
             dependencies=[Depends(require_menu_access("hris_attendance"))])
def create_attendance_manual(
    request:      Request,
    payload:      AttendanceManualCreate,
    current_user: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    emp = db.query(Employee).filter(Employee.id == payload.employee_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")

    existing = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == payload.employee_id,
        AttendanceRecord.date        == payload.date,
    ).first()
    if existing:
        raise HTTPException(409, "Attendance record already exists for this date. Use PATCH to update.")

    # Auto-calculate hours if clock_in and clock_out provided
    reg = payload.hours_regular
    ot_wd = payload.hours_overtime_weekday
    ot_we = payload.hours_overtime_weekend
    ot_hol = payload.hours_overtime_holiday

    if payload.clock_in and payload.clock_out and reg is None:
        weekend = _is_weekend(payload.date)
        reg, ot_wd, ot_we, ot_hol = _calculate_overtime(
            payload.clock_in, payload.clock_out, weekend, _is_holiday(db, payload.date)
        )

    record = AttendanceRecord(
        employee_id            = payload.employee_id,
        date                   = payload.date,
        clock_in               = payload.clock_in,
        clock_out              = payload.clock_out,
        hours_regular          = reg,
        hours_overtime_weekday = ot_wd,
        hours_overtime_weekend = ot_we,
        hours_overtime_holiday = ot_hol,
        source                 = AttendanceSource.MANUAL,
        note                   = payload.note,
        face_verified          = False,
    )
    db.add(record)
    db.flush()
    linked_overtime_ids = _link_overtime_requests_to_attendance(db, record)
    write_audit(db, "AttendanceRecord", record.id, "MANUAL_CREATE",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                after={**model_to_dict(record),
                       "linked_overtime_request_ids": linked_overtime_ids})
    db.commit()
    db.refresh(record)
    return record


# ─── Attendance: list ─────────────────────────────────────────────────────────

@router.get("/attendance", response_model=PaginatedResponse[AttendanceRecordResponse],
            summary="List attendance records",
            dependencies=[Depends(require_menu_access("hris_attendance"))])
def list_attendance(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    employee_id:  int | None = Query(None),
    date_from:    date | None = Query(None),
    date_to:      date | None = Query(None),
    work_group_id: int | None = Query(None),
    skip:         int         = Query(0, ge=0),
    limit:        int         = Query(50, ge=1, le=200),
):
    q = db.query(AttendanceRecord)

    visible_ids = _visible_employee_ids(db, current_user)
    if visible_ids is not None:
        if not visible_ids:
            return {"items": [], "total": 0}
        q = q.filter(AttendanceRecord.employee_id.in_(visible_ids))
    if employee_id:
        q = q.filter(AttendanceRecord.employee_id == employee_id)

    if work_group_id:
        q = q.join(Employee, Employee.id == AttendanceRecord.employee_id).filter(
            Employee.work_group_id == work_group_id
        )

    if date_from:
        q = q.filter(AttendanceRecord.date >= date_from)
    if date_to:
        q = q.filter(AttendanceRecord.date <= date_to)

    total = q.count()
    items = q.order_by(AttendanceRecord.date.desc()).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


# ─── Attendance: monthly summary ─────────────────────────────────────────────

@router.get(
    "/attendance/summary",
    summary="Monthly attendance summary per employee",
    dependencies=[Depends(require_menu_access("hris_attendance"))],
)
def attendance_summary(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    year:         int = Query(...),
    month:        int = Query(...),
    dept_id:      int | None = Query(None),
):
    from calendar import monthrange
    first_day = date(year, month, 1)
    last_day  = date(year, month, monthrange(year, month)[1])

    q = (
        db.query(
            AttendanceRecord.employee_id,
            func.count(AttendanceRecord.clock_in).label("days_present"),
            func.coalesce(func.sum(AttendanceRecord.hours_regular),          0).label("hours_regular"),
            func.coalesce(func.sum(AttendanceRecord.hours_overtime_weekday),  0).label("hours_ot_weekday"),
            func.coalesce(func.sum(AttendanceRecord.hours_overtime_weekend),  0).label("hours_ot_weekend"),
            func.coalesce(func.sum(AttendanceRecord.hours_overtime_holiday),  0).label("hours_ot_holiday"),
        )
        .join(Employee, Employee.id == AttendanceRecord.employee_id)
        .filter(
            AttendanceRecord.date >= first_day,
            AttendanceRecord.date <= last_day,
        )
        .group_by(AttendanceRecord.employee_id)
    )

    visible_ids = _visible_employee_ids(db, current_user)
    if visible_ids is not None:
        if not visible_ids:
            return []
        q = q.filter(AttendanceRecord.employee_id.in_(visible_ids))
    if dept_id:
        q = q.filter(Employee.dept_id == dept_id)

    rows = q.all()
    emp_ids = [r.employee_id for r in rows]
    emp_map = {
        e.id: e for e in db.query(Employee).filter(Employee.id.in_(emp_ids)).all()
    }

    result = []
    for r in rows:
        emp = emp_map.get(r.employee_id)
        if not emp:
            continue
        hours_ot_weekday = Decimal(str(r.hours_ot_weekday))
        hours_ot_weekend = Decimal(str(r.hours_ot_weekend))
        hours_ot_holiday = Decimal(str(r.hours_ot_holiday))
        hours_ot_total = hours_ot_weekday + hours_ot_weekend + hours_ot_holiday
        hours_regular = Decimal(str(r.hours_regular))
        result.append({
            "employee_id":    r.employee_id,
            "employee_no":    emp.employee_no,
            "full_name":      emp.full_name,
            "department":     emp.department.name if emp.department else None,
            "days_present":   r.days_present,
            "hours_regular":  str(hours_regular),
            "hours_overtime_weekday": str(hours_ot_weekday),
            "hours_overtime_weekend": str(hours_ot_weekend),
            "hours_overtime_holiday": str(hours_ot_holiday),
            "hours_ot_total": str(hours_ot_total),
            "total_hours": str(hours_regular + hours_ot_total),
        })

    return result


# ─── Attendance: export (Excel/CSV) ─────────────────────────────────────────

@router.get(
    "/attendance/export",
    summary="Export attendance records as Excel",
    dependencies=[Depends(require_menu_access("hris_attendance"))],
)
def export_attendance(
    current_user: Annotated[CurrentUser, Depends(require_role(
        RoleName.SUPER_ADMIN, RoleName.MD, RoleName.PM, RoleName.GA, RoleName.FINANCE, RoleName.COST_CONTROL
    ))],
    db:           Annotated[Session, Depends(get_db)],
    date_from:    date | None = Query(None),
    date_to:      date | None = Query(None),
    dept_id:      int | None  = Query(None),
    employee_id:  int | None  = Query(None),
    fmt:          str         = Query("xlsx", pattern="^(xlsx|csv)$"),
    timezone_offset_minutes: int = Query(
        0,
        ge=MIN_BROWSER_TIMEZONE_OFFSET,
        le=MAX_BROWSER_TIMEZONE_OFFSET,
    ),
):
    """Download attendance records as Excel (.xlsx) or CSV."""
    q = (
        db.query(AttendanceRecord)
        .join(Employee, Employee.id == AttendanceRecord.employee_id)
    )
    visible_ids = _visible_employee_ids(db, current_user)
    if visible_ids is not None:
        if not visible_ids:
            q = q.filter(False)
        else:
            q = q.filter(AttendanceRecord.employee_id.in_(visible_ids))
    if date_from:
        q = q.filter(AttendanceRecord.date >= date_from)
    if date_to:
        q = q.filter(AttendanceRecord.date <= date_to)
    if dept_id:
        q = q.filter(Employee.dept_id == dept_id)
    if employee_id:
        q = q.filter(AttendanceRecord.employee_id == employee_id)

    records = q.order_by(Employee.full_name, AttendanceRecord.date).all()

    # Build employee lookup
    emp_ids = {r.employee_id for r in records}
    emp_map = {e.id: e for e in db.query(Employee).filter(Employee.id.in_(emp_ids)).all()}
    wl_map: dict[int, WorkLocation] = {}
    wl_ids = {e.work_location_id for e in emp_map.values() if e.work_location_id}
    if wl_ids:
        wl_map = {w.id: w for w in db.query(WorkLocation).filter(WorkLocation.id.in_(wl_ids)).all()}

    HEADERS = [
        "Tanggal", "No. Karyawan", "Nama", "Departemen", "Lokasi Kerja",
        "Jam Masuk", "Jam Keluar",
        "Jam Reguler", "OT Weekday", "OT Weekend", "OT Libur",
        "Total Jam OT", "Sumber",
        "Latitude", "Longitude", "Akurasi (m)",
        "Lokasi OK", "Jarak (m)",
        "Wajah Terdeteksi", "Catatan",
    ]

    client_timezone = timezone_from_browser_offset(timezone_offset_minutes)

    def _fmt_time(dt: datetime | None) -> str:
        if dt is None:
            return ""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(client_timezone).strftime("%H:%M:%S")

    def _fmt_dec(v) -> str:
        return str(round(float(v), 2)) if v is not None else ""

    rows = []
    for r in records:
        emp = emp_map.get(r.employee_id)
        dept_name = emp.department.name if emp and emp.department else ""
        wl = wl_map.get(emp.work_location_id) if emp and emp.work_location_id else None
        ot_total = sum(
            float(x or 0) for x in [
                r.hours_overtime_weekday,
                r.hours_overtime_weekend,
                r.hours_overtime_holiday,
            ]
        )
        rows.append([
            str(r.date),
            emp.employee_no if emp else "",
            emp.full_name if emp else "",
            dept_name,
            wl.name if wl else "",
            _fmt_time(r.clock_in),
            _fmt_time(r.clock_out),
            _fmt_dec(r.hours_regular),
            _fmt_dec(r.hours_overtime_weekday),
            _fmt_dec(r.hours_overtime_weekend),
            _fmt_dec(r.hours_overtime_holiday),
            str(round(ot_total, 2)),
            r.source.value if r.source else "",
            _fmt_dec(r.latitude),
            _fmt_dec(r.longitude),
            _fmt_dec(r.accuracy),
            "Ya" if r.location_ok is True else ("Tidak" if r.location_ok is False else ""),
            _fmt_dec(r.location_distance_m),
            "Ya" if r.face_verified else "Tidak",
            r.note or "",
        ])

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        writer.writerows(rows)
        output.seek(0)
        fname = f"attendance_{date_from or 'all'}_{date_to or 'all'}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — use fmt=csv")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Absensi"

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-fit column widths (approximation)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"attendance_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── Work Locations CRUD ─────────────────────────────────────────────────────

_wl_roles = (RoleName.SUPER_ADMIN, RoleName.MD, RoleName.PM, RoleName.PROJECT_CONTROL, RoleName.GA, RoleName.HR)


@router.get("/work-locations", response_model=list[WorkLocationResponse],
            summary="List work locations",
            dependencies=[Depends(require_menu_access("hris_attendance", "hris_settings"))])
def list_work_locations(
    _:           CurrentUser,
    db:          Annotated[Session, Depends(get_db)],
    active_only: bool = Query(True),
):
    q = db.query(WorkLocation)
    if active_only:
        q = q.filter(WorkLocation.is_active == True)
    return q.order_by(WorkLocation.name).all()


@router.post("/work-locations", response_model=WorkLocationResponse, status_code=201,
             summary="Create a work location",
             dependencies=[Depends(require_menu_access("hris_settings"))])
def create_work_location(
    payload:      WorkLocationCreate,
    current_user: Annotated[CurrentUser, Depends(require_role(*_wl_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    wl = WorkLocation(**payload.model_dump())
    db.add(wl)
    db.commit()
    db.refresh(wl)
    return wl


@router.patch("/work-locations/{wl_id}", response_model=WorkLocationResponse,
              summary="Update a work location",
              dependencies=[Depends(require_menu_access("hris_settings"))])
def update_work_location(
    wl_id:        int,
    payload:      WorkLocationUpdate,
    current_user: Annotated[CurrentUser, Depends(require_role(*_wl_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    wl = db.query(WorkLocation).filter(WorkLocation.id == wl_id).first()
    if not wl:
        raise HTTPException(404, "Work location not found")
    for field, val in payload.model_dump(exclude_unset=True).items():
        setattr(wl, field, val)
    db.commit()
    db.refresh(wl)
    return wl


@router.patch("/employees/{employee_id}/work-location",
              response_model=dict,
              summary="Assign or clear work location for an employee",
              dependencies=[Depends(require_menu_access("hris_employees", "hris_settings"))])
def assign_employee_work_location(
    employee_id:      int,
    current_user:     Annotated[CurrentUser, Depends(require_role(*_wl_roles))],
    db:               Annotated[Session, Depends(get_db)],
    work_location_id: int | None = Query(None, description="Pass null to clear assignment"),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")

    if work_location_id is not None:
        wl = db.query(WorkLocation).filter(WorkLocation.id == work_location_id).first()
        if not wl:
            raise HTTPException(404, "Work location not found")
        emp.work_location_id = wl.id
        msg = f"Assigned to {wl.name}"
    else:
        emp.work_location_id = None
        msg = "Work location cleared"

    db.commit()
    return {"message": msg, "employee_id": employee_id, "work_location_id": work_location_id}


# ─── Leave Types ─────────────────────────────────────────────────────────────

@router.get(
    "/leave-types",
    response_model=list[LeaveTypeResponse],
    summary="List leave types",
    dependencies=[Depends(require_menu_access("hris_leave", "hris_settings"))],
)
def list_leave_types(
    _:  CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    active_only: bool = True,
):
    q = db.query(LeaveType)
    if active_only:
        q = q.filter(LeaveType.is_active == True)
    return q.order_by(LeaveType.id).all()


@router.post("/leave-types", response_model=LeaveTypeResponse, status_code=201,
             summary="Create leave type",
             dependencies=[Depends(require_menu_access("hris_settings"))])
def create_leave_type(
    payload:      LeaveTypeCreate,
    current_user: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    if db.query(LeaveType).filter(LeaveType.code == payload.code).first():
        raise HTTPException(409, "Leave type code already exists")
    lt = LeaveType(**payload.model_dump())
    db.add(lt)
    db.commit()
    db.refresh(lt)
    return lt


# ─── Leave Balance ────────────────────────────────────────────────────────────

@router.get("/leave-balance/{employee_id}", response_model=list[LeaveBalanceResponse],
            summary="Get leave balances for employee (current year)",
            dependencies=[Depends(require_menu_access("hris_leave"))])
def get_leave_balance(
    employee_id:  int,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    year:         int = Query(default=None),
):
    if year is None:
        year = datetime.now(timezone.utc).year

    # Ensure employee exists
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")

    visible_ids = _visible_employee_ids(db, current_user)
    if visible_ids is not None and employee_id not in visible_ids:
        raise HTTPException(403, "Access denied")

    balances = (
        db.query(LeaveBalance)
        .filter(LeaveBalance.employee_id == employee_id, LeaveBalance.year == year)
        .all()
    )
    return balances


@router.post(
    "/leave-balance/seed",
    summary="Seed leave balances for all active employees (HR admin)",
    dependencies=[Depends(require_menu_access("hris_leave", "hris_settings"))],
)
def seed_leave_balances(
    current_user: Annotated[CurrentUser, Depends(require_role(*_hr_roles, RoleName.MD))],
    db:           Annotated[Session, Depends(get_db)],
    year:         int = Query(default=None),
):
    """Ensure every active employee has a balance row for each active leave type."""
    if year is None:
        year = datetime.now(timezone.utc).year

    employees = db.query(Employee).filter(
        Employee.status.in_(["active", "probation"])
    ).all()
    leave_types = db.query(LeaveType).filter(LeaveType.is_active == True).all()
    created = 0

    if not employees or not leave_types:
        return MessageResponse(message=f"Seeded {created} leave balance rows for {year}")

    emp_ids = [e.id for e in employees]
    lt_ids  = [lt.id for lt in leave_types]

    # Load all existing balances for this year in ONE query (N+1 fix)
    existing_keys: set[tuple[int, int]] = {
        (b.employee_id, b.leave_type_id)
        for b in db.query(LeaveBalance).filter(
            LeaveBalance.employee_id.in_(emp_ids),
            LeaveBalance.leave_type_id.in_(lt_ids),
            LeaveBalance.year == year,
        ).all()
    }

    lt_map = {lt.id: lt for lt in leave_types}
    for emp in employees:
        for lt in leave_types:
            if (emp.id, lt.id) not in existing_keys:
                db.add(LeaveBalance(
                    employee_id   = emp.id,
                    leave_type_id = lt.id,
                    year          = year,
                    accrued       = lt.max_days_per_year or 0,
                    used          = 0,
                ))
                created += 1

    db.commit()
    return MessageResponse(message=f"Seeded {created} leave balance rows for {year}")


# ─── Leave Requests ───────────────────────────────────────────────────────────

@router.get(
    "/leave-requests/duration-preview",
    summary="Preview leave working days",
    dependencies=[Depends(require_menu_access("hris_leave"))],
)
def preview_leave_duration(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    start_date:   date = Query(...),
    end_date:     date = Query(...),
):
    days, excluded_holidays = _leave_duration(db, start_date, end_date)
    return {"days": days, "excluded_holidays": excluded_holidays}


@router.post("/leave-requests/doctor-certificate", status_code=201,
             summary="Upload a medical certificate for a leave request",
             dependencies=[Depends(require_menu_access("hris_leave"))])
async def upload_leave_doctor_certificate(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    file:         UploadFile = File(...),
):
    _cleanup_orphan_leave_certificates(db, current_user.id)
    ext = Path(file.filename or "").suffix.lower()
    if ext not in _LEAVE_CERT_EXTENSIONS or file.content_type not in _LEAVE_CERT_CONTENT_TYPES:
        raise HTTPException(400, "Doctor certificate must be PDF, JPEG, or PNG")
    content = await file.read()
    if not content:
        raise HTTPException(400, "Doctor certificate file is empty")
    if len(content) > _MAX_LEAVE_CERT_BYTES:
        raise HTTPException(413, "Doctor certificate exceeds the 10 MB limit")

    filename = f"user_{current_user.id}_{uuid.uuid4().hex}{ext}"
    (_LEAVE_CERT_DIR / filename).write_bytes(content)
    return {"file_url": f"/uploads/leave_certificates/{filename}"}


@router.delete(
    "/leave-requests/doctor-certificate",
    response_model=MessageResponse,
    summary="Discard an unsubmitted medical certificate",
    dependencies=[Depends(require_menu_access("hris_leave"))],
)
def discard_leave_doctor_certificate(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    file_url:     str = Query(...),
):
    expected_prefix = "/uploads/leave_certificates/"
    if not file_url.startswith(expected_prefix):
        raise HTTPException(422, "Invalid doctor certificate URL")
    filename = Path(file_url).name
    if not filename.startswith(f"user_{current_user.id}_"):
        raise HTTPException(403, "Doctor certificate belongs to another user")
    if db.query(LeaveRequest.id).filter(LeaveRequest.doctor_cert_url == file_url).first():
        raise HTTPException(409, "Doctor certificate is already attached to a leave request")
    path = (_LEAVE_CERT_DIR / filename).resolve()
    if not path.is_relative_to(_LEAVE_CERT_DIR.resolve()):
        raise HTTPException(403, "Access denied")
    path.unlink(missing_ok=True)
    return MessageResponse(message="Unsubmitted doctor certificate discarded")


@router.get("/leave-requests", response_model=PaginatedResponse[LeaveRequestResponse],
            summary="List leave requests",
            dependencies=[Depends(require_menu_access("hris_leave"))])
def list_leave_requests(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    employee_id:  int | None                = Query(None),
    req_status:   LeaveRequestStatus | None = Query(None, alias="status"),
    skip:         int                       = Query(0, ge=0),
    limit:        int                       = Query(50, ge=1, le=200),
):
    q = db.query(LeaveRequest)

    visible_ids = _visible_employee_ids(db, current_user)
    if visible_ids is not None:
        if not visible_ids:
            return {"items": [], "total": 0}
        q = q.filter(LeaveRequest.employee_id.in_(visible_ids))
    if employee_id:
        q = q.filter(LeaveRequest.employee_id == employee_id)

    if req_status:
        q = q.filter(LeaveRequest.status == req_status)

    total = q.count()
    items = q.order_by(LeaveRequest.id.desc()).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


@router.post("/leave-requests", response_model=LeaveRequestResponse, status_code=201,
             summary="Submit a leave request",
             dependencies=[Depends(require_menu_access("hris_leave"))])
def submit_leave_request(
    request:      Request,
    payload:      LeaveRequestCreate,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    if payload.employee_id:
        emp = db.query(Employee).filter(Employee.id == payload.employee_id).first()
        if not emp:
            raise HTTPException(404, "Employee not found")
        # HR/admin can submit for others; employees can only submit for themselves
        if emp.user_id != current_user.id and current_user.role.name not in _hr_roles:
            raise HTTPException(403, "Can only submit leave for yourself")
    else:
        emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not emp:
            raise HTTPException(404, "No employee record linked to your account")
    ensure_employee_can_use_self_service(emp)

    lt = db.query(LeaveType).filter(
        LeaveType.id == payload.leave_type_id, LeaveType.is_active == True
    ).first()
    if not lt:
        raise HTTPException(404, "Leave type not found or inactive")

    # Calculate business days, excluding weekends and configured holidays.
    delta, _ = _leave_duration(db, payload.start_date, payload.end_date)
    if delta == 0:
        raise HTTPException(422, "Leave dates must include at least one working day")

    if lt.requires_doctor_cert:
        if not payload.doctor_cert_url:
            raise HTTPException(422, "Doctor certificate is required for this leave type")
        if not payload.doctor_cert_url.startswith("/uploads/leave_certificates/"):
            raise HTTPException(422, "Invalid doctor certificate URL")
        cert_name = Path(payload.doctor_cert_url).name
        cert_path = (_LEAVE_CERT_DIR / cert_name).resolve()
        if (
            not cert_path.is_relative_to(_LEAVE_CERT_DIR.resolve())
            or not cert_path.is_file()
            or not cert_name.startswith(f"user_{current_user.id}_")
        ):
            raise HTTPException(422, "Doctor certificate is missing or belongs to another user")
        if db.query(LeaveRequest).filter(
            LeaveRequest.doctor_cert_url == payload.doctor_cert_url
        ).first():
            raise HTTPException(409, "Doctor certificate is already attached to another request")

    overlap = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id == emp.id,
        LeaveRequest.status.in_([
            LeaveRequestStatus.SUBMITTED,
            LeaveRequestStatus.APPROVED,
        ]),
        LeaveRequest.start_date <= payload.end_date,
        LeaveRequest.end_date >= payload.start_date,
    ).first()
    if overlap:
        raise HTTPException(409, f"Leave dates overlap request #{overlap.id}")

    allocations = _leave_days_by_year(db, payload.start_date, payload.end_date)
    _check_or_deduct_leave_balances(db, emp.id, lt, allocations)

    # Build approval chain
    chain = _LEAVE_APPROVAL_CHAIN if lt.requires_approval else []

    req = LeaveRequest(
        employee_id           = emp.id,
        leave_type_id         = lt.id,
        start_date            = payload.start_date,
        end_date              = payload.end_date,
        days                  = delta,
        reason                = payload.reason,
        doctor_cert_url       = payload.doctor_cert_url,
        status                = LeaveRequestStatus.SUBMITTED if chain else LeaveRequestStatus.APPROVED,
        approval_chain        = chain,
        approval_step         = 0,
        current_approver_role = chain[0] if chain else None,
        approval_history      = [{
            "action": "SUBMIT",
            "role": None,
            "user_id": current_user.id,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "note": None,
        }],
        submitted_by = current_user.id,
    )
    db.add(req)
    db.flush()

    # If auto-approved (no approval required), deduct balance
    if req.status == LeaveRequestStatus.APPROVED:
        _check_or_deduct_leave_balances(db, emp.id, lt, allocations, deduct=True)

    write_audit(db, "LeaveRequest", req.id, "SUBMIT",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                after=model_to_dict(req))
    db.commit()
    db.refresh(req)

    # Notify approver
    if chain:
        push_to_role(db, RoleName[chain[0]],
                     "Pengajuan Cuti Baru",
                     f"{emp.full_name} mengajukan cuti {lt.name} {delta} hari",
                     "/hris/leave")
        db.commit()

    return req


@router.post("/leave-requests/{req_id}/approve", response_model=LeaveRequestResponse,
             summary="Approve a leave request",
             dependencies=[Depends(require_menu_access("hris_leave"))])
def approve_leave_request(
    req_id:       int,
    request:      Request,
    payload:      LeaveActionRequest,
    current_user: Annotated[CurrentUser, Depends(require_role(*_mgr_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    req = _get_leave_or_404(req_id, db, lock=True)

    if req.status != LeaveRequestStatus.SUBMITTED:
        raise HTTPException(409, f"Cannot approve: status is '{req.status.value}'")

    _require_current_leave_approver(req, current_user)

    chain = req.approval_chain or []
    step  = req.approval_step + 1

    _add_leave_history(req, current_user.id, "APPROVE", payload.note)

    if step >= len(chain):
        # Final approval
        req.status                = LeaveRequestStatus.APPROVED
        req.current_approver_role = None
        req.approved_by           = current_user.id
        req.approval_step         = step
        # Lock and recheck balances at final approval to prevent concurrent overspend.
        allocations = _leave_days_by_year(db, req.start_date, req.end_date)
        _check_or_deduct_leave_balances(
            db, req.employee_id, req.leave_type, allocations, deduct=True,
        )
        # Notify employee
        if req.employee and req.employee.user_id:
            push(db, req.employee.user_id,
                 "Cuti Disetujui",
                 f"Pengajuan cuti {req.leave_type.name} {req.days} hari telah disetujui",
                 "/hris/me/leave")
    else:
        # Advance to next approver
        req.approval_step         = step
        req.current_approver_role = chain[step]
        push_to_role(
            db,
            RoleName(chain[step]),
            "Pengajuan Cuti Menunggu Persetujuan",
            f"Pengajuan cuti {(req.employee.full_name if req.employee else f'#{req.employee_id}')} "
            "menunggu persetujuan Anda",
            "/hris/leave",
        )

    write_audit(db, "LeaveRequest", req.id, "APPROVE",
                changed_by=current_user.id, ip_address=get_client_ip(request))
    db.commit()
    db.refresh(req)
    return req


@router.post("/leave-requests/{req_id}/reject", response_model=LeaveRequestResponse,
             summary="Reject a leave request",
             dependencies=[Depends(require_menu_access("hris_leave"))])
def reject_leave_request(
    req_id:       int,
    request:      Request,
    payload:      LeaveActionRequest,
    current_user: Annotated[CurrentUser, Depends(require_role(*_mgr_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    req = _get_leave_or_404(req_id, db, lock=True)

    if req.status != LeaveRequestStatus.SUBMITTED:
        raise HTTPException(409, f"Cannot reject: status is '{req.status.value}'")

    _require_current_leave_approver(req, current_user)
    _add_leave_history(req, current_user.id, "REJECT", payload.note)
    req.status = LeaveRequestStatus.REJECTED

    write_audit(db, "LeaveRequest", req.id, "REJECT",
                changed_by=current_user.id, ip_address=get_client_ip(request))
    db.commit()
    db.refresh(req)

    # Notify employee
    if req.employee and req.employee.user_id:
        push(db, req.employee.user_id,
             "Cuti Ditolak",
             f"Pengajuan cuti {req.leave_type.name} {req.days} hari ditolak. {payload.note or ''}",
             "/hris/me/leave")
        db.commit()

    return req


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _get_leave_or_404(req_id: int, db: Session, lock: bool = False) -> LeaveRequest:
    query = db.query(LeaveRequest).filter(LeaveRequest.id == req_id)
    if lock:
        query = query.with_for_update()
    req = query.first()
    if not req:
        raise HTTPException(404, "Leave request not found")
    return req


def _add_leave_history(req: LeaveRequest, actor_id: int, action: str, note: str | None):
    history = list(req.approval_history or [])
    history.append({
        "action":    action,
        "role":      req.current_approver_role,
        "user_id":   actor_id,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "note":      note,
    })
    req.approval_history = history


# ═══════════════════════════════════════════════════════════════════════════════
# Overtime Requests
# ═══════════════════════════════════════════════════════════════════════════════

def _get_my_employee_att(db: Session, cu, *, lock: bool = False) -> Employee:
    query = db.query(Employee).filter(Employee.user_id == cu.id)
    if lock:
        query = query.with_for_update()
    emp = query.first()
    if not emp:
        raise HTTPException(404, "No employee profile linked to your account")
    return ensure_employee_can_use_self_service(emp)


@router.post("/overtime-requests", response_model=OvertimeRequestResponse, status_code=201,
             summary="Submit overtime request",
             dependencies=[Depends(require_menu_access("hris_attendance"))])
def submit_overtime_request(
    request: Request,
    payload: OvertimeRequestCreate,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    emp = _get_my_employee_att(db, cu, lock=True)
    existing = (
        db.query(OvertimeRequest)
        .filter(
            OvertimeRequest.employee_id == emp.id,
            OvertimeRequest.date == payload.date,
            OvertimeRequest.status.in_([
                OvertimeRequestStatus.SUBMITTED,
                OvertimeRequestStatus.APPROVED,
            ]),
        )
        .first()
    )
    if existing:
        raise HTTPException(409, "An active overtime request already exists for this date")

    attendance = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.employee_id == emp.id,
            AttendanceRecord.date == payload.date,
        )
        .first()
    )
    ot = OvertimeRequest(
        employee_id=emp.id,
        date=payload.date,
        planned_hours=payload.planned_hours,
        reason=payload.reason,
        status=OvertimeRequestStatus.SUBMITTED,
        attendance_id=attendance.id if attendance else None,
    )
    db.add(ot)
    db.flush()
    write_audit(
        db, "OvertimeRequest", ot.id, "SUBMIT",
        changed_by=cu.id, ip_address=get_client_ip(request),
        after=model_to_dict(ot),
    )
    push_to_role(db, RoleName.GA,
                 "Pengajuan Lembur Baru",
                 f"{emp.full_name} mengajukan lembur {payload.planned_hours}j pada {payload.date}",
                 "/hris/attendance")
    db.commit()
    db.refresh(ot)
    resp = OvertimeRequestResponse.model_validate(ot)
    resp.employee_name = emp.full_name
    return resp


@router.get("/me/overtime-requests", response_model=list[OvertimeRequestResponse],
            summary="My overtime requests",
            dependencies=[Depends(require_menu_access("hris_attendance"))])
def my_overtime_requests(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    emp = _get_my_employee_att(db, cu)
    rows = (
        db.query(OvertimeRequest)
        .filter(OvertimeRequest.employee_id == emp.id)
        .order_by(OvertimeRequest.date.desc())
        .all()
    )
    result = []
    for r in rows:
        resp = OvertimeRequestResponse.model_validate(r)
        resp.employee_name = emp.full_name
        result.append(resp)
    return result


@router.get("/overtime-requests", response_model=list[OvertimeRequestResponse],
            summary="List all overtime requests (HR/MD)",
            dependencies=[Depends(require_menu_access("hris_attendance"))])
def list_overtime_requests(
    cu: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db: Annotated[Session, Depends(get_db)],
    status_filter: str | None = Query(None, alias="status"),
    date_from: date | None = Query(None),
    date_to:   date | None = Query(None),
):
    q = db.query(OvertimeRequest)
    if status_filter:
        q = q.filter(OvertimeRequest.status == status_filter)
    if date_from:
        q = q.filter(OvertimeRequest.date >= date_from)
    if date_to:
        q = q.filter(OvertimeRequest.date <= date_to)
    rows = q.order_by(OvertimeRequest.date.desc()).all()
    result = []
    for r in rows:
        resp = OvertimeRequestResponse.model_validate(r)
        resp.employee_name = r.employee.full_name if r.employee else None
        result.append(resp)
    return result


@router.post(
    "/overtime-requests/{ot_id}/approve",
    response_model=OvertimeRequestResponse,
    dependencies=[Depends(require_menu_access("hris_attendance"))],
)
def approve_overtime_request(
    ot_id: int,
    request: Request,
    payload: OvertimeActionRequest,
    cu: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db: Annotated[Session, Depends(get_db)],
):
    ot = (
        db.query(OvertimeRequest)
        .filter(OvertimeRequest.id == ot_id)
        .with_for_update()
        .first()
    )
    if not ot:
        raise HTTPException(404, "Overtime request not found")
    if ot.status != OvertimeRequestStatus.SUBMITTED:
        raise HTTPException(400, f"Request already {ot.status.value}")
    before = model_to_dict(ot)
    ot.status = OvertimeRequestStatus.APPROVED
    ot.approved_by = cu.id
    ot.approved_at = datetime.now(timezone.utc)
    if ot.attendance_id is None:
        attendance = (
            db.query(AttendanceRecord)
            .filter(
                AttendanceRecord.employee_id == ot.employee_id,
                AttendanceRecord.date == ot.date,
            )
            .first()
        )
        ot.attendance_id = attendance.id if attendance else None
    write_audit(
        db, "OvertimeRequest", ot.id, "APPROVE",
        changed_by=cu.id, ip_address=get_client_ip(request),
        before=before, after=model_to_dict(ot),
    )
    if ot.employee and ot.employee.user_id:
        push(db, ot.employee.user_id, "Lembur Disetujui",
             f"Pengajuan lembur Anda pada {ot.date} telah disetujui.",
             "/hris/me/overtime")
    db.commit()
    db.refresh(ot)
    resp = OvertimeRequestResponse.model_validate(ot)
    resp.employee_name = ot.employee.full_name if ot.employee else None
    return resp


@router.post(
    "/overtime-requests/{ot_id}/reject",
    response_model=OvertimeRequestResponse,
    dependencies=[Depends(require_menu_access("hris_attendance"))],
)
def reject_overtime_request(
    ot_id: int,
    request: Request,
    payload: OvertimeActionRequest,
    cu: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db: Annotated[Session, Depends(get_db)],
):
    ot = (
        db.query(OvertimeRequest)
        .filter(OvertimeRequest.id == ot_id)
        .with_for_update()
        .first()
    )
    if not ot:
        raise HTTPException(404, "Overtime request not found")
    if ot.status != OvertimeRequestStatus.SUBMITTED:
        raise HTTPException(400, f"Request already {ot.status.value}")
    before = model_to_dict(ot)
    ot.status = OvertimeRequestStatus.REJECTED
    ot.approved_by = cu.id
    ot.approved_at = datetime.now(timezone.utc)
    ot.rejection_reason = payload.note
    write_audit(
        db, "OvertimeRequest", ot.id, "REJECT",
        changed_by=cu.id, ip_address=get_client_ip(request),
        before=before, after=model_to_dict(ot),
    )
    if ot.employee and ot.employee.user_id:
        push(db, ot.employee.user_id, "Lembur Ditolak",
             f"Pengajuan lembur Anda pada {ot.date} ditolak. {payload.note or ''}",
             "/hris/me/overtime")
    db.commit()
    db.refresh(ot)
    resp = OvertimeRequestResponse.model_validate(ot)
    resp.employee_name = ot.employee.full_name if ot.employee else None
    return resp


# ═══════════════════════════════════════════════════════════════════════════════
# Team Leave Calendar
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/leave-requests/calendar", response_model=list[LeaveCalendarItem],
            summary="Team leave calendar for a given month",
            dependencies=[Depends(require_menu_access("hris_leave"))])
def leave_calendar(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    year:    int = Query(default=None),
    month:   int = Query(default=None),
    dept_id: int | None = Query(None),
):
    today = date.today()
    year  = year  or today.year
    month = month or today.month

    month_start = date(year, month, 1)
    # Last day of month
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)

    q = (
        db.query(LeaveRequest)
        .join(Employee, LeaveRequest.employee_id == Employee.id)
        .filter(
            LeaveRequest.status == LeaveRequestStatus.APPROVED,
            LeaveRequest.end_date >= month_start,
            LeaveRequest.start_date <= month_end,
        )
    )
    visible_ids = _visible_employee_ids(db, cu)
    if visible_ids is not None:
        if not visible_ids:
            return []
        q = q.filter(LeaveRequest.employee_id.in_(visible_ids))
    if dept_id:
        q = q.filter(Employee.dept_id == dept_id)

    rows = q.all()
    result = []
    for r in rows:
        result.append(LeaveCalendarItem(
            employee_id=r.employee_id,
            employee_name=r.employee.full_name if r.employee else "Unknown",
            dept=r.employee.department.name if r.employee and r.employee.department else None,
            leave_type=r.leave_type.name if r.leave_type else "—",
            start_date=r.start_date,
            end_date=r.end_date,
            days=r.days,
            status=r.status.value,
        ))
    result.sort(key=lambda x: x.start_date)
    return result
